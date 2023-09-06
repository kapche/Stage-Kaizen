"""
:author KAPCHE TEBU BAUDOUIN
:version 1.0
"""

import verification
import openpyxl  # Module openpyxl pour la manipulation de fichiers Excel.
import creation
import suppression
import extraction
import os
import re  # Module pour les expressions régulières (regex).
import main

appel=main.appel
suppres_pdf=main.suppres_pdf
candidat=main.candidat
init_candidat=main.init_candidat
verif = main.verif
note_garde=main.note_garde

def ajout_appel(fichier):
    """
    Cette fonction gère l'ajout de fichiers d'appel à partir du répertoire "appel" vers le répertoire "candidat".
    """
    global appel
    global suppres_pdf
    global candidat
    global init_candidat
    global verif 
    
    ajout = appel + "/" + fichier
    fichier_ajoute = verification.verification_extend(ajout)
    creation.nouvel_fiche(fichier_ajoute)
    
    if suppres_pdf:
        suppression.suppression_pdf(fichier_ajoute)
        
    fiche = suppression.suppression_extend(fichier)
    postulant = candidat + "/" + fiche
    init_candidat[fiche] = extraction.obtenir_fichiers_dans_dossier(postulant)
    verif = False

def ajout_de_donnes(besoin, competence, profil, nom, excel):
    """
    Cette fonction ajoute des données à un fichier Excel existant.

    :param besoin: Les besoins du projet.
    :type besoin: str
    :param competence: Les compétences du profil.
    :type competence: str
    :param profil: Les informations du profil.
    :type profil: list
    :param nom: Le nom du candidat ou du projet.
    :type nom: str
    :param excel: Le chemin du fichier Excel où ajouter les données.
    :type excel: str
    """
    while True:
        try:
            
            classeur = openpyxl.load_workbook(excel)

            feuille = classeur["projet"]
            metier = classeur["metier"]

            global verif
            verif = True

            for i in range(7, 2, -2):
                index = 0
                if verif == False:
                    break
                verification.verification_mot(profil, i, feuille, index, excel, competence, besoin, classeur, nom)
            print(verif)

            if verif == True:
                projet = ["NEX'US", "Eric Mercandali", "", "", "", "", profil[0]]
                projet_ligne = feuille.max_row + 1

                for col, valeur in enumerate(projet, start=1):
                    feuille.cell(row=projet_ligne, column=col, value=valeur)
                classeur.save(excel)

                ajout_metier(projet_ligne, excel, competence, besoin, profil, classeur, feuille, nom)
                
            classeur.close()
            break
        except Exception as e:
            print(f"Une erreur s'est produite au niveau de l'ajout de donnes: {str(e)}")

def ajout_metier(index, nom, competence, besoin, profil, classeur, feuille, dossier):
    """
    Cette fonction ajoute des données à la feuille "metier" d'un fichier Excel.

    :param index: L'index de la ligne dans la feuille "projet" correspondant au profil.
    :type index: int
    :param nom: Le nom du fichier Excel où ajouter les données.
    :type nom: str
    :param competence: Les compétences du profil.
    :type competence: list
    :param besoin: Les besoins du projet.
    :type besoin: list
    :param profil: Les informations du profil.
    :type profil: list
    :param classeur: Le classeur Excel.
    :type classeur: openpyxl.Workbook
    :param feuille: La feuille "projet" du classeur Excel.
    :type feuille: openpyxl.Worksheet
    :param dossier: Le nom du dossier du chemin.
    :type dossier: str
    """
    metier = classeur["metier"]

    # Obtenez les valeurs des cellules de la feuille "projet" pour le profil donné
    cellule1 = feuille["A" + str(index)]
    cellule2 = feuille["C" + str(index)]
    cellule3 = feuille["E" + str(index)]
    cellule4 = feuille["G" + str(index)]
    cellule5 = profil[2] + profil[3] 
    cellule6 = ', '.join(competence)
    cellule7 = ', '.join(besoin)
    cellule8 = dossier.split("/")[1]

    new_ligne = [cellule1.value, cellule2.value, cellule3.value, cellule4.value, cellule5, cellule6, cellule7, cellule8]

    # Trouvez la dernière ligne vide dans la feuille "metier"
    derniere_ligne = metier.max_row + 1

    # Ajoutez les données à la nouvelle ligne
    for col, valeur in enumerate(new_ligne, start=1):
        metier.cell(row=derniere_ligne, column=col, value=valeur)

    classeur.save(nom)

def ajout_cv(fichier,dossier):
    """
    Cette fonction gère l'ajout d'un CV à un dossier, effectue des vérifications et extrait des informations du CV pour les stocker dans un fichier Excel.
    """
    global note_garde
    global suppres_pdf
    
    fichier_ajoute = dossier + "/" + fichier
    fichier_ajoute = verification.verification_extend(fichier_ajoute)
    note_cv(fichier_ajoute, main.excel)
    if suppres_pdf:
        fichier_sans_dossier = fichier_ajoute.replace(dossier + "/", "")
        note_garde.append(fichier_sans_dossier)
        suppres_pdf=False

def note_cv(CV, excel):
    """
    Cette fonction note un candidat en fonction des compétences extraites de son CV et enregistre les informations
    dans un fichier Excel spécifié.

    :param dossier: Le chemin du dossier du candidat contenant le CV.
    :type dossier: str
    :param excel: Le chemin du fichier Excel où enregistrer les informations de notation.
    :type excel: str
    """
    if os.path.exists(CV):
        
        # Récupération du contenu du CV 
        cv = extraction.extract_cv(CV)
        text = ""
        for item in cv:
            text += item + " "

        # Utilisation d'expressions régulières pour extraire le numéro de téléphone et l'email du texte du CV
        pattern = r'\b0[1-9](?:[-.\s]?\d{2}){4}\b'
        matches = re.findall(pattern, text)
        numero = ','.join(matches) 

        pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
        matches = re.findall(pattern, text)
        email = ','.join(matches) 

        global candidat
        
        # Obtenir des informations sur le candidat à partir du chemin du dossier
        info = CV.replace(candidat, "").split("/")[1]
        identite = suppression.suppression_extend(CV.replace(candidat, "").split("/")[-1])

        while True:
            try:
                # Charger le classeur Excel
                classeur = openpyxl.load_workbook(excel)

                # Sélectionnez la feuille sur laquelle vous souhaitez travailler
                metier = classeur["metier"]
                notation = classeur["notation"]

                global verif
                verif = True
                index = 0

                # Recherchez la ligne correspondante dans la feuille "metier" en fonction des informations du candidat
                for ligne in metier.iter_rows(min_col=8, max_col=8, values_only=True):
                    if verif == False:
                        break
                    index += 1
                    for cellule in ligne:
                        if verif == False:
                            break
                        if cellule is not None and isinstance(cellule, str) and info.lower() in cellule.lower():
                            verif = False

                # Récupérer d'autres informations sur le candidat à partir de la ligne correspondante dans la feuille "metier"
                comptence_primaire = metier["F" + str(index)].value
                pole = metier["A" + str(index)].value
                departement = metier["B" + str(index)].value
                projet = metier["C" + str(index)].value
                application = metier["D" + str(index)].value
                travail = metier["E" + str(index)].value

                if comptence_primaire == None:
                    competence = []
                else:
                    competence = comptence_primaire.split(", ")
                print(competence)

                niveau = 0

                # Calcul du niveau de compétence du candidat en fonction des compétences requises
                for case in competence:
                    for item in cv:
                        if item.lower().find(case.lower()) != -1:
                            niveau += 1

                # Créez une nouvelle entrée pour le candidat dans la feuille "notation"
                new_candidat = [pole, departement, projet, application, travail, identite, numero, email, niveau]
                derniere_ligne = notation.max_row + 1

                # Ajoutez les données à la nouvelle ligne dans la feuille "notation"
                for col, valeur in enumerate(new_candidat, start=1):
                    notation.cell(row=derniere_ligne, column=col, value=valeur)
                classeur.save(excel)
                classeur.close()
                break
            except Exception as e:
                print(f"Une erreur s'est produite au niveau de note: {str(e)}")