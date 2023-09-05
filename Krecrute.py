#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os  # Module pour interagir avec le système de fichiers et les chemins d'accès.
import win32com.client  # Module pour utiliser COM (Component Object Model) sur les systèmes Windows.
from pdfminer.high_level import extract_text  # Module pdfminer pour extraire le texte à partir de fichiers PDF.
import io  # Module pour travailler avec des flux d'octets (bytes streams).
import pdfplumber  # Module pdfplumber pour extraire des données à partir de fichiers PDF.
import numpy as np  # Module pour effectuer des calculs scientifiques, principalement utilisé pour les tableaux.
import pandas as pd  # Module pour la manipulation et l'analyse de données tabulaires.
import openpyxl  # Module openpyxl pour la manipulation de fichiers Excel.
import time  # Module pour gérer le temps et les retards dans le code.
import re  # Module pour les expressions régulières (regex).
import shutil  # Module pour la manipulation de fichiers et de répertoires.
import subprocess  # Module pour exécuter des commandes système.
from openpyxl import Workbook  # Module openpyxl pour la manipulation de fichiers Excel.
from openpyxl.utils.dataframe import dataframe_to_rows  # Fonction pour convertir des données de DataFrame en lignes Excel.
import pythoncom  # Module pour gérer la communication avec COM (Component Object Model) sur Windows.
import aspose.words as aw  # Module Aspose.Words pour la manipulation de documents Word.

"""
La fonction creation(dossier) vérifie si le dossier spécifié existe déjà. Si le dossier n'existe pas, la fonction crée le dossier en utilisant os.makedirs(dossier) et affiche un message indiquant que le dossier a été créé. Cette fonction est utile lorsque vous avez besoin de vous assurer qu'un dossier existe avant de continuer à travailler avec des fichiers à l'intérieur. Elle facilite également la gestion des dossiers dans votre code en évitant des erreurs liées à l'absence de dossiers nécessaires.
"""
# In[2]:


def creation(dossier):
    if not os.path.exists(dossier):
        os.makedirs(dossier)
        print(f"Le dossier '{dossier}' a été créé.")

"""
La fonction extract(pdf_path) prend en entrée le chemin d'accès d'un fichier PDF (pdf_path) et utilise la bibliothèque pdfminer pour extraire le texte contenu dans ce fichier. Elle commence par ouvrir le fichier PDF en mode lecture binaire ('rb'), puis elle utilise pdfminer pour extraire le texte. Le texte extrait est ensuite divisé en lignes en utilisant le caractère de saut de ligne ('\n'), et ces lignes sont stockées dans une liste appelée lines. Enfin, la fonction renvoie la liste lines contenant le texte extrait du PDF. Cette fonction est utile pour l'extraction de texte à partir de fichiers PDF, ce qui peut être utile dans diverses applications de traitement de texte et d'analyse de données.
"""
# In[3]:


def extract(pdf_path):
    lines = []
    with open(pdf_path, 'rb') as pdf_file:
    # Utiliser pdfminer pour extraire le texte
        text = extract_text(io.BytesIO(pdf_file.read()))
        lines.extend(text.strip().split('\n'))
    return lines

"""
La fonction extract_appel(pdf_path) utilise la bibliothèque pdfplumber pour extraire le texte d'un fichier PDF donné par son chemin d'accès. Elle parcourt chaque page du PDF, extrait le texte de chaque page, le divise en lignes, et stocke ces lignes dans une liste appelée lines. En fin de compte, la fonction renvoie cette liste contenant tout le texte extrait du PDF. Cela permet une extraction efficace du texte à partir de documents PDF, en particulier lorsqu'il s'agit de fichiers PDF multi-pages.
"""
# In[4]:


def extract_appel(pdf_path):
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            lines.extend(text.strip().split('\n'))
    return lines


# In[5]:


# Fonction pour extraire des compétences à partir d'un fichier Excel
def extract_competence():
    while True:  # Boucle infinie pour réessayer en cas d'erreur
        try:
            global excel  # Utilisation de la variable globale 'excel' qui contient le chemin du fichier Excel
            competence = []  # Une liste pour stocker les compétences extraites

            # Charger le fichier Excel
            classeur = openpyxl.load_workbook(excel)

            # Liste des noms de feuilles que vous souhaitez lire
            feuilles = ["competence"]

            # Parcourir la liste des feuilles
            for feuille_nom in feuilles:
                feuille = classeur[feuille_nom]

                # Parcourir les lignes de la feuille
                for ligne in feuille.iter_rows():
                    for cellule in ligne:
                        competence.append(cellule.value)  # Ajouter la valeur de chaque cellule à la liste 'competence'

            # Fermer le fichier Excel
            classeur.close()

            return competence  # Renvoyer la liste des compétences extraites

        except Exception as e:
            creation_xlsx("krecrute.xlsx")  # Créer un fichier Excel 'krecrute.xlsx' en cas d'erreur
            print(f"Une erreur s'est produite au niveau d'extract_competence: {str(e)}")


# In[6]:


# Fonction pour traiter un nouveau document PDF
def nouvel_fiche(new):
    # Extraire le texte du document PDF en utilisant extract_appel
    lines = extract_appel(new)

    # Définir les chaînes de recherche pour les sections pertinentes du document
    search_string1 = "EXPRESSION DE BESOIN"
    search_string2 = "CONDITIONS D'EXÉCUTION DE LA PRESTATION"
    
    # Initialiser des variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in lines:
        if found == True:
            text += lines[i]
        if item == search_string1:
            found = True
        if item == search_string2:
            found = False
        i += 1

    # Diviser le texte extrait en éléments de besoins
    besoin = text.split(".")
    del besoin[-1]
    
    # Définir de nouvelles chaînes de recherche pour d'autres sections pertinentes
    search_string1 = "Consultation"
    search_string2 = "Description du projet :"
    
    # Réinitialiser les variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in lines:
        if found == True:
            text += (lines[i] + "\n")
        if item.find(search_string1) != -1:
            found = True
        if item == search_string2:
            found = False
        i += 1
    
    # Diviser le texte extrait en éléments de profil
    profil1 = text.split("\n")
    del profil1[-1]
    del profil1[-1]
    
    # Vérifier et ajuster si nécessaire la première ligne du profil
    if profil1[0].find("Projet") == -1:
        del profil1[0]
    
    # Diviser les éléments de profil en intitulés et descriptifs
    pro = []
    for item in profil1:
        pro += item.split(":")
    
    intitule_profil = []
    descriptif_profil = []
    i = 0
    for item in pro:
        if i % 2 == 0:
            intitule_profil.append(item)
        else:
            descriptif_profil.append(item)
        i += 1

    # Définir de nouvelles chaînes de recherche pour d'autres sections pertinentes
    search_string1 = "Stack technique du projet :"
    search_string2 = "Compétences attendues sur le profil :"
    
    # Réinitialiser les variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in lines:
        if found == True:
            text += (lines[i] + "\n")
        if item.find(search_string1) != -1:
            found = True
        if item.find(search_string2) != -1:
            found = False
        i += 1

    # Diviser le texte extrait en éléments de stack technique
    stack = text.split("\n")
    del stack[-1]
    del stack[-1]

    # Extraire les compétences de la base de données
    competence1 = extract_competence()

    # Créer une liste pour stocker les compétences du profil
    competence_profil = []
    
    # Comparer les compétences extraites avec le texte du stack technique
    for case in competence1:
        for item in stack:
            if re.search(case, item, re.IGNORECASE):
                if case not in competence_profil:
                    competence_profil.append(case)

    # Supprimer les extensions du nom du fichier
    new = suppression_extend(new)
    
    # Extraire le nom du fichier en supprimant la partie de l'appel
    nom = new.replace(appel, "")

    # Afficher les besoins, le descriptif du profil et le nom
    print(besoin)
    print("\n")
    print(descriptif_profil)

    # Ajouter les données extraites au fichier Excel global
    global excel
    ajout_de_donnes(besoin, competence_profil, descriptif_profil, nom, excel)

    # Créer un répertoire pour le candidat et le projet
    url = candidat + "/" + nom
    creation(url)

    print("Initialisation complète de la nouvelle fiche")


# In[7]:


# Fonction pour vérifier les fichiers dans un dossier
def verification(dossier, init):
    # Obtenir la liste des fichiers présents dans le dossier
    fichiers_apres = obtenir_fichiers_dans_dossier(dossier)
    
    # Comparer les listes pour détecter les ajouts et suppressions
    fichiers_ajoutes = [fichier for fichier in fichiers_apres if fichier not in init]
    fichiers_supprimes = [fichier for fichier in init if fichier not in fichiers_apres]
    
    # Afficher les fichiers ajoutés et supprimés
    print("Fichiers ajoutés:", fichiers_ajoutes)
    print("Fichiers supprimés:", fichiers_supprimes)
    
    # Renvoyer la liste des fichiers actuels dans le dossier
    return fichiers_apres    


# In[8]:


# Fonction pour supprimer un dossier et son contenu
def suppression(dossier):
    global candidat  # Utilisation de la variable globale 'candidat' qui spécifie le répertoire racine des candidats
    global init_candidat  # Utilisation de la variable globale 'init_candidat' qui contient des informations initiales
    
    # Parcourir les éléments initiaux du dossier donné
    for item in init_candidat[dossier]:
        item = suppression_extend(item)  # Appel d'une fonction non définie (peut nécessiter une implémentation)
        suppression_note(item)  # Appel d'une fonction non définie (peut nécessiter une implémentation)
    
    # Construire le chemin complet du dossier
    fiche = candidat + "/" + dossier
    
    # Supprimer le dossier s'il existe
    if os.path.exists(fiche) and os.path.isdir(fiche):
        shutil.rmtree(fiche)
        print(f"Le dossier '{fiche}' a été supprimé.")
    else:
        print(f"Le dossier '{fiche}' n'existe pas.")
    
    print("Suppression terminée")


# In[9]:


# Fonction pour noter un candidat et enregistrer les informations dans un fichier Excel
def note(dossier, excel):
    # Vérifiez si le dossier du candidat existe
    if os.path.exists(dossier):
        # Récupération du contenu du CV depuis le dossier (supposons que la fonction extract() récupère le texte du CV)
        cv = extract(dossier)
        text = ""
        for item in cv:
            text += item + " "

        # Utilisation d'expressions régulières pour extraire le numéro de téléphone et l'email du texte du CV
        pattern = r'\b0[1-9](?:[-.\s]?\d{2}){4}\b'
        matches = re.findall(pattern, text)
        numero = ','.join(matches)  # Fusionner les numéros s'il y en a plusieurs
        print(numero)

        pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
        matches = re.findall(pattern, text)
        email = ','.join(matches)  # Fusionner les emails s'il y en a plusieurs
        print(email)

        global candidat
        # Obtenir des informations sur le candidat à partir du chemin du dossier
        info = dossier.replace(candidat, "").split("/")[1]
        identite = suppression_extend(dossier.replace(candidat, "").split("/")[-1])
        print(info)

        while True:
            try:
                # Charger le classeur Excel
                classeur = openpyxl.load_workbook(excel)

                # Sélectionnez la feuille sur laquelle vous souhaitez travailler
                metier = classeur["metier"]
                notation = classeur["notation"]

                global verif
                verif = True
                index = 0

                # Recherchez la ligne correspondante dans la feuille "metier" en fonction des informations du candidat
                for ligne in metier.iter_rows(min_col=8, max_col=8, values_only=True):
                    if verif == False:
                        break
                    index += 1
                    for cellule in ligne:
                        if verif == False:
                            break
                        if cellule is not None and isinstance(cellule, str) and info.lower() in cellule.lower():
                            # Faites votre vérification ici, par exemple, imprimez la valeur de la cellule
                            print(f"Valeur de la cellule : '{cellule}' à la position {index}")
                            verif = False

                # Récupérer d'autres informations sur le candidat à partir de la ligne correspondante dans la feuille "metier"
                compte = metier["F" + str(index)].value
                pole = metier["A" + str(index)].value
                departement = metier["B" + str(index)].value
                projet = metier["C" + str(index)].value
                application = metier["D" + str(index)].value
                taf = metier["E" + str(index)].value

                if compte == None:
                    competence = []
                else:
                    competence = compte.split(", ")
                print(competence)

                niveau = 0

                # Calcul du niveau de compétence du candidat en fonction des compétences requises
                for case in competence:
                    for item in cv:
                        if item.lower().find(case.lower()) != -1:
                            niveau += 1
                print(niveau)

                # Créez une nouvelle entrée pour le candidat dans la feuille "notation"
                new_candidat = [pole, departement, projet, application, taf, identite, numero, email, niveau]
                derniere_ligne = notation.max_row + 1

                # Ajoutez les données à la nouvelle ligne dans la feuille "notation"
                for col, valeur in enumerate(new_candidat, start=1):
                    notation.cell(row=derniere_ligne, column=col, value=valeur)
                classeur.save(excel)
                classeur.close()
                break
            except Exception as e:
                print(f"Une erreur s'est produite au niveau de note: {str(e)}")


# In[10]:


# Fonction pour supprimer des notes d'un fichier Excel en fonction du nom de fichier
def suppression_note(fichier):
    while True:
        try:
            global excel  # Utilisation de la variable globale 'excel' qui contient le chemin du fichier Excel
            
            # Charger le fichier Excel
            classeur = openpyxl.load_workbook(excel)
            print(fichier)
            
            # Sélectionner la feuille de travail "notation"
            feuille = classeur['notation']
            
            index = 0
            supp = True
            
            # Parcourir les lignes de la colonne 6 (F) dans la feuille "notation"
            for ligne in feuille.iter_rows(min_col=6, max_col=6, values_only=True):
                if supp == False:
                    break
                index += 1
                for cellule in ligne:
                    if supp == False:
                        break
                    if cellule is not None and isinstance(cellule, str) and fichier.lower() in cellule.lower():
                        # Faites votre vérification ici, par exemple, imprimez la valeur de la cellule
                        print(f"Valeur de la cellule : '{cellule}' à la position {index}")
                        
                        # Supprimer la ligne correspondante dans la feuille "notation"
                        feuille.delete_rows(index)

            # Enregistrer les modifications dans le fichier Excel
            classeur.save(excel)
            
            # Fermer le fichier Excel
            classeur.close()
            
            print("Note supprimée avec succès")
            break
        except Exception as e:
            print(f"Une erreur s'est produite au niveau de la suppression de note : {str(e)}")


# In[11]:


# Fonction pour créer un nouveau fichier Excel avec des feuilles de travail
def creation_xlsx(nom):
    excel = krecrute + "/" + nom  # Définir le chemin complet du nouveau fichier Excel
    
    if not os.path.exists(excel):
        # Créez un nouveau classeur Excel
        classeur = openpyxl.Workbook()
        
        # Supprimez la feuille de travail par défaut "Sheet"
        nom_de_la_feuille_a_supprimer = "Sheet"
        if nom_de_la_feuille_a_supprimer in classeur.sheetnames:
            feuille_a_supprimer = classeur[nom_de_la_feuille_a_supprimer]
            classeur.remove(feuille_a_supprimer)

        # Créez les feuilles de travail requises
        feuille_projet = classeur.create_sheet("projet")
        feuille_competence = classeur.create_sheet("competence")
        feuille_notation = classeur.create_sheet("notation")
        feuille_metier = classeur.create_sheet("metier")
        feuille_contact = classeur.create_sheet("contact")
        
        # Définissez des données pour les compétences et les projets (ces données sont fournies dans le code)
        # Liste des compétences
        competences=["java","angular","scala","spark","spring","hadoop","postgresql","jenkins","kubernetes","ansible","java EE",
                     "Architecture microservice","spring boot","react","react js","html","git","gitlab","maven","virtualbox",
                     "vagrant","docker","tls","mongodb","sonar","junit","protractor","java spring boot","wildfly","weblogic",
                     "angular js","css","kafka","helm","gitlab ci","cdi","jax-rs","jms","shell","oracle","systeme unix","rhel",
                     "pki","protocol tls","grafana","influxDB","prometheus","collectd","telegraf","appDymanics","jmxtrans",
                     "micrometer","sql devellopper","soapui","rest","telco","cosem","lambda expressions","EJB Runtimes",
                     "terraforme","spring core","JPA","intel IJ","cucumber","kalka","kotlin","even driven architecture",
                     "postgres sql","freeraduis","jdfc","logflash","elasticsearch","telegral","filebeat","spring cloud stream",
                     "linux"]
        
        # Dictionnaire des projets
        projet={'pole':["NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US",
                        "NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US",
                        "NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US"],
                'responsable de pole':["Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali"],
                'departement':["SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC",
                               "CCMM","CCMM","CCMM","CCMM","CCMM","CCMM","CMMA","CMMA","CMMA","CMMA","CMMA","CMMA","SEAL",
                               "SEAL","SEAL","SEAL","SEAL"],
                'responsable de departement':["Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Cedric GAUDET","Cedric GAUDET","Cedric GAUDET","Cedric GAUDET","Cedric GAUDET",
                                              "Cedric GAUDET","pierre GOTELAERE","pierre GOTELAERE","pierre GOTELAERE",
                                              "pierre GOTELAERE","pierre GOTELAERE","pierre GOTELAERE","Eric PERRIER",
                                              "Eric PERRIER","Eric PERRIER","Eric PERRIER","Eric PERRIER"],
                'projet':["ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","mobilite electricque",
                          "mobilite electricque","mobilite electricque","mobilite electricque","MAGELLAN","Gestion Clients",
                          "Gestion Clients","Gestion Clients","usine de Dév","RTE","LINKY ET MAGELLAN","recette transverse",
                          "POA","BEM","deploiment IP","techonologie 2G/3G","PACCMAN","PACCMAN","PACCMAN","iMAT","Expertises",
                          "Expertises","Expertises","Expertises","Cybersecurité"],
                'responsable de projet':["Jean-Bernard Lepidi","Jean-Bernard Lepidi","Jean-Bernard Lepidi",
                                         "Jean-Bernard Lepidi","Jean-Bernard Lepidi","Jean-Bernard Lepidi",
                                         "Jean-Bernard Lepidi","Jean-Bernard Lepidi","Vincent Delalande","Vincent Delalande",
                                         "Vincent Delalande","Vincent Delalande","Alandji BOUORNAKA","Abdourrahmane CHRIABI",
                                         "Abdourrahmane CHRIABI","Abdourrahmane CHRIABI","Sophie CLAPIER","olivier LEROUX",
                                         "cyril Marcant","chrystelle ilicinkas","jean luc LACHAUD","","","","Vincent Boucheux",
                                         "Vincent Boucheux","Vincent Boucheux","Jean-Marc LEBRETON","Benjamin BRETAULT",
                                         "Benjamin BRETAULT","Benjamin BRETAULT","Benjamin BRETAULT","Marc NAUDIN"],
                'application':["CARTOLINE","CAP'ten","Diag BT","PIX'F","PIX'R","Saturne","Tableau de bord energetique",
                               "Bilan metrologiques","mobilite interne","smart charging","comptages du futur","etudes DATA","",
                               "Appui avant vente","realisation e projets smart metting",
                               "appui a la gestion des contrats des clinets externes","","","","","","","","","ICSS","SAR",
                               "Icoeur","","Nouveau comptage et veille technologique","TELCO WAN","OTC","Appui au pilotage",
                               "SEAL Securite"],
                'autre responsable':["","","","","","","","","","","","","","","","","PASCALE ROBERT","PASCALE ROBERT",
                                     "PASCALE ROBERT","PASCALE ROBERT","PASCALE ROBERT","PASCALE ROBERT",
                                     "Anne-Christine LECLERQ","Anne-Christine LECLERQ","Anne-Christine LECLERQ",
                                     "Anne-Christine LECLERQ","Anne-Christine LECLERQ","Anne-Christine LECLERQ",
                                     "RACHED BOUDAOUA","RACHED BOUDAOUA","RACHED BOUDAOUA","RACHED BOUDAOUA",
                                     "RACHED BOUDAOUA"]
               }
        
        # Dictionnaire des contacts (noms et e-mails)
        contact ={"nom":["Nicolas Pigeon","Cedric GAUDET","pierre GOTELAERE","Eric PERRIER","Jean-Bernard Lepidi",
                         "Vincent Delalande","Alandji BOUORNAKA","Abdourrahmane CHRIABI","Sophie CLAPIER","olivier LEROUX",
                         "cyril Marcant","chrystelle ilicinkas","jean luc LACHAUD","Jean-Marc LEBRETON","Benjamin BRETAULT",
                         "Marc NAUDIN","PASCALE ROBERT","Anne-Christine LECLERQ","RACHED BOUDAOUA","Vincent Boucheux"],
                  "email":["nicolas.pigeon@enedis.fr","cedric.gaudet@enedis.fr","pierre.gotelaire@enedis.fr",
                           "eric.perrier@enedis.fr","jean-bernard.lepidi@enedis.fr","vincent.delalande@enedis.fr",
                           "alandji.bouornaka@enedis.fr","abdourrahmane.chriabi@enedis.fr","sophie.clapier@enedis.fr",
                           "olivier.leroux@enedis.fr","cyril.marcant@enedis.fr","chrystelle.ilicinkas@enedis.fr",
                           "jean.lachaud@enedis.fr","jean-marc.lebreton@enedis.fr","benjamin.bretault@enedis.fr",
                           "marc.naudin@enedis.fr","pascale.robert@enedis.fr","anne-christine.leclerq@enedis.fr",
                          "rached.boudaoua@enedis.fr","vincent.boucheux@enedis.fr"],
                  "contact":["","","","","","","","","","","","","","","","","","","",""]
            }

         # Dictionnaire des données métier
        metier={"pole":[],"departement":[],"projet":[],"application":[],"metier":[],"competence":[],"besoin":[],
                "identifiant":[]}
        
        # Dictionnaire des données de notation
        notation={"pole":[],"departement":[],"projet":[],"application":[],"metier":[],"identifiants":[],"contact":[],"email":[],
                  "note":[]}

        # Créez des DataFrames pandas à partir des données
        df=pd.DataFrame(projet)
        df1=pd.DataFrame(contact)
        df2=pd.DataFrame(metier)
        df3=pd.DataFrame(notation)

       # Choisissez la colonne dans laquelle vous souhaitez ajouter les données (par exemple, colonne A)
        colonne = "A"

        # Utilisez une boucle pour ajouter les données des compétences à la feuille "competence"
        for index, donnee in enumerate(competences, start=1):
            feuille_competence[f"{colonne}{index}"] = donnee

        # Utilisez dataframe_to_rows pour ajouter les données des DataFrames pandas aux feuilles de travail correspondantes
        for row in dataframe_to_rows(df, index=False, header=True):
            feuille_projet.append(row)

        for row in dataframe_to_rows(df1, index=False, header=True):
            feuille_contact.append(row)

        for row in dataframe_to_rows(df2, index=False, header=True):
            feuille_metier.append(row)

        for row in dataframe_to_rows(df3, index=False, header=True):
            feuille_notation.append(row)

        # Sauvegardez le classeur dans un fichier Excel
        classeur.save(excel)

        # Fermez le classeur
        classeur.close()


# In[12]:


# Fonction pour obtenir la liste des fichiers dans un dossier
def obtenir_fichiers_dans_dossier(dossier):
    return [fichier for fichier in os.listdir(dossier)]


# In[13]:


def chemin():
    # Exécutez la commande pour obtenir le chemin du répertoire "Documents"
    command = r'reg query "HKEY_CURRENT_USER\Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders" /v Personal'
    result = subprocess.run(command, stdout=subprocess.PIPE, shell=True, text=True)

    # Analysez la sortie de la commande pour obtenir le chemin du répertoire "Documents"
    lines = result.stdout.strip().split('\n')
    documents_path_line = lines[-1]
    _, _, documents_path = documents_path_line.partition("    ")

    print("Chemin du répertoire Documents:", documents_path)
    position = documents_path.split("    ")  # Sépare la ligne en une liste ["Type", "REG_SZ", "Chemin"]
    print(position)
    chemin = position[-1].replace("\\", "/")  # Récupère le chemin et remplace les barres obliques inverses par des barres obliques normales
    if chemin.find("%USERPROFILE%") != -1:
        chemin = chemin.replace("/Documents", "")  # Si le chemin contient "%USERPROFILE%", retire "/Documents"
    return chemin


# In[14]:


def suppression_pdf(fiche):
    global suppres_pdf  # Utilisation d'une variable globale suppres_pdf

    # Vérifier si le fichier existe avant de le supprimer
    if os.path.exists(fiche):
        os.remove(fiche)  # Supprimer le fichier PDF
        suppres_pdf = False  # Mettre la variable globale suppres_pdf à False pour indiquer que la suppression a été effectuée
        print(f"Le fichier {fiche} a été supprimé avec succès.")


# In[15]:


def suppression_extend(fiche):
    # Extraire l'extension du fichier en séparant le nom de fichier à l'aide de os.path.splitext()
    extension = os.path.splitext(fiche)[0]

    return extension  # Renvoyer la partie du nom de fichier avant l'extension


# In[16]:


def verification_extend(fiche):
    extension = os.path.splitext(fiche)  # Obtenir l'extension du fichier
    global suppres_pdf

    # Vérifier si l'extension est .docx ou .doc
    if extension[1] == ".docx" or extension[1] == ".doc":
        # Spécifiez le chemin complet de votre fichier .doc
        fichier_doc = fiche.replace('/', "\\\\")  # Remplacez les barres obliques par des doubles barres obliques

        # Obtenez le chemin du fichier PDF de sortie (remplacez l'extension .doc par .pdf)
        fichier_pdf = os.path.splitext(fichier_doc)[0] + ".pdf"

        while True:
            try:
                # Utilisez la bibliothèque Aspose.Words pour convertir le fichier .doc/.docx en PDF
                doc = aw.Document(fichier_doc)
                doc.save(fichier_pdf)
                break  # Sortez de la boucle une fois la conversion réussie

            except Exception as e:
                print(f"Une erreur s'est produite au niveau de la verification d'extention : {str(e)}")

        suppres_pdf = True  # Indique que la conversion a été effectuée avec succès
        name = list(extension)
        name[1] = ".pdf"
        fiche = ""
        for item in name:
            fiche += item  # Construisez le nom de fichier PDF converti

    return fiche  # Renvoie le nom de fichier d'origine ou le nom du fichier PDF converti


# In[17]:


def ajout_de_donnes(besoin, competence, profil, nom, excel):
    while True:
        try:
            # Ouvrez le fichier Excel
            classeur = openpyxl.load_workbook(excel)

            # Sélectionnez la feuille sur laquelle vous souhaitez travailler
            feuille = classeur["projet"]
            metier = classeur["metier"]

            global verif
            verif = True

            # Parcours des lignes de la feuille Excel à partir de la ligne 7 jusqu'à la ligne 2 avec un pas de 2
            for i in range(7, 2, -2):
                index = 0
                if verif == False:
                    break
                # Appelez la fonction verification_mot pour vérifier les informations du profil
                verification_mot(profil, i, feuille, index, excel, competence, besoin, classeur, nom)
            print(verif)

            # Si verif est toujours True (aucun profil existant ne correspond), ajoutez le profil
            if verif == True:
                projet = ["NEX'US", "Eric Mercandali", "", "", "", "", profil[0]]
                projet_ligne = feuille.max_row + 1

                # Ajoutez les données à la nouvelle ligne
                for col, valeur in enumerate(projet, start=1):
                    feuille.cell(row=projet_ligne, column=col, value=valeur)
                classeur.save(excel)

                # Appelez la fonction ajout_metier pour ajouter les informations aux feuilles "metier" et "competence"
                ajout_metier(projet_ligne, excel, competence, besoin, profil, classeur, feuille, nom)

            # Fermez le classeur Excel
            classeur.close()
            break
        except Exception as e:
            print(f"Une erreur s'est produite au niveau de l'ajout de donnes: {str(e)}")


# In[18]:


def ajout_metier(index, nom, competence, besoin, profil, classeur, feuille, dossier):
    metier = classeur["metier"]

    # Obtenez les valeurs des cellules de la feuille "projet" pour le profil donné
    cellule1 = feuille["A" + str(index)]
    cellule2 = feuille["C" + str(index)]
    cellule3 = feuille["E" + str(index)]
    cellule4 = feuille["G" + str(index)]
    cellule5 = profil[2] + profil[3]  # Concaténez les éléments du profil
    cellule6 = ', '.join(competence)  # Concaténez les compétences avec des virgules
    cellule7 = ', '.join(besoin)  # Concaténez les besoins avec des virgules
    cellule8 = dossier.split("/")[1]  # Obtenez le nom du dossier du chemin

    # Créez une nouvelle ligne avec les données
    new_ligne = [cellule1.value, cellule2.value, cellule3.value, cellule4.value, cellule5, cellule6, cellule7, cellule8]

    # Trouvez la dernière ligne vide dans la feuille "metier"
    derniere_ligne = metier.max_row + 1

    # Ajoutez les données à la nouvelle ligne
    for col, valeur in enumerate(new_ligne, start=1):
        metier.cell(row=derniere_ligne, column=col, value=valeur)

    # Sauvegardez le classeur Excel
    classeur.save(nom)


# In[19]:


def verification_mot(profil, indexe, feuille, index, excel, competence, besoin, classeur, nom):
    global verif  # Utilisation d'une variable globale pour contrôler la vérification
    for ligne in feuille.iter_rows(min_col=indexe, max_col=indexe, values_only=True):
        if verif == False:  # Si la vérification a déjà été effectuée, sortir de la boucle
            break
        index += 1
        for cellule in ligne:
            if verif == False:  # Si la vérification a déjà été effectuée, sortir de la boucle
                break
            if cellule is not None and isinstance(cellule, str) and profil[0].replace(" ", "").lower() in cellule.lower():
                # Vérifiez si le nom du profil (en minuscules et sans espaces) est dans la cellule
                print(f"Valeur de la cellule : '{cellule}' à la position {index}")
                # Le nom du profil a été trouvé dans la cellule, appel à la fonction ajout_metier()
                ajout_metier(index, excel, competence, besoin, profil, classeur, feuille, nom)
                verif = False  # La vérification est terminée, ne pas continuer


# In[20]:


def verification_instance():
    # Déclaration des variables globales
    global krecrute
    global appel
    global candidat
    global excel
    global init_appel
    global init_candidat
    global init_dossier
    
    # Vérification et création du répertoire "krecrute" s'il n'existe pas
    if not os.path.exists(krecrute):
        creation(krecrute)
    
    # Vérification et création du répertoire "appel" s'il n'existe pas
    if not os.path.exists(appel):
        # Si le répertoire "candidat" existe et est un dossier, supprimez son contenu
        if os.path.exists(candidat) and os.path.isdir(candidat):
            # Parcourez tous les fichiers et dossiers à l'intérieur du dossier "candidat"
            for element in os.listdir(candidat):
                element_path = os.path.join(candidat, element)

                # Vérifiez s'il s'agit d'un fichier
                if os.path.isfile(element_path):
                    # Supprimez le fichier
                    os.remove(element_path)
                elif os.path.isdir(element_path):
                    # Supprimez le dossier (et son contenu récursivement)
                    os.rmdir(element_path)
        
        # Créez le répertoire "appel"
        creation(appel)
        init_appel = []  # Réinitialisez la liste des fichiers d'appel
        init_candidat = {}  # Réinitialisez le dictionnaire des fichiers candidats
        init_dossier = []  # Réinitialisez la liste des dossiers
    
    # Vérification et création du répertoire "candidat" s'il n'existe pas
    if not os.path.exists(candidat):
        # Si le répertoire "candidat" existe et est un dossier, supprimez son contenu
        for cle, valeur in init_candidat.items():
            for item in init_candidat[cle]:
                item = suppression_extend(item)
                suppression_note(item)
        
        # Créez le répertoire "candidat"
        creation(candidat)
        init_appel = []  # Réinitialisez la liste des fichiers d'appel
        init_candidat = {}  # Réinitialisez le dictionnaire des fichiers candidats
        init_dossier = []  # Réinitialisez la liste des dossiers
    
    # Vérification et création du fichier Excel "excel" s'il n'existe pas
    if not os.path.exists(excel):
        # Créez un nouveau fichier Excel "krecrute.xlsx"
        creation_xlsx("krecrute.xlsx")
        init_appel = []  # Réinitialisez la liste des fichiers d'appel
        init_candidat = {}  # Réinitialisez le dictionnaire des fichiers candidats
        init_dossier = []  # Réinitialisez la liste des dossiers
    
    # Vérification et création des dossiers dans le répertoire "candidat"
    for cle, valeur in init_candidat.items():
        dossier = candidat + "/" + cle
        print(dossier)
        if not os.path.exists(dossier):
            for item in init_candidat[cle]:
                item = suppression_extend(item)
                suppression_note(item)
            
            # Créez le dossier pour chaque clé du dictionnaire
            creation(dossier)
            init_candidat[cle] = []  # Réinitialisez la liste des fichiers candidats pour ce dossier


# In[21]:


def ajout_appel(fichier):
    # Déclaration des variables globales
    global appel
    global suppres_pdf
    global candidat
    global init_candidat
    global verif 
    
    # Créez le chemin complet du fichier dans le répertoire "appel"
    ajout = appel + "/" + fichier
    
    # Utilisez la fonction "verification_extend" pour vérifier et convertir l'extension du fichier si nécessaire
    ajout = verification_extend(ajout)
    
    # Appelez la fonction pour créer le nouveau fichier dans le répertoire "appel"
    nouvel_fiche(ajout)
    
    # Si le drapeau "suppres_pdf" est activé, supprimez le fichier PDF
    if suppres_pdf:
        suppression_pdf(ajout)
    
    # Obtenez le nom du fichier sans extension
    fiche = suppression_extend(fichier)
    
    # Créez le chemin complet du fichier dans le répertoire "candidat"
    chercheur = candidat + "/" + fiche
    
    # Obtenez la liste des fichiers dans le dossier "chercheur" et stockez-la dans le dictionnaire "init_candidat"
    init_candidat[fiche] = obtenir_fichiers_dans_dossier(chercheur)
    
    # Réinitialisez le drapeau "verif" à True
    verif = True


# In[22]:


def suppression_appel(fichier):
    # Déclaration de la variable globale
    global init_candidat
    
    # Obtenez le nom du fichier sans extension
    fiche = suppression_extend(fichier)
    
    # Appelez la fonction "suppression" pour supprimer le dossier correspondant au fichier
    suppression(fiche)
    
    # Supprimez également l'entrée correspondante dans le dictionnaire "init_candidat"
    del init_candidat[fiche]


# In[23]:


def ajout_cv(fichier):
    # Déclaration des variables globales
    global dossier
    global note_garde
    global suppres_pdf
    
    # Construisez le chemin complet du fichier
    chemin_complet = dossier + "/" + fichier
    
    # Vérifiez et convertissez le fichier si nécessaire en PDF et extrayez des informations
    choix = verification_extend(chemin_complet)
    
    # Appeler la fonction "note" pour extraire des informations à partir du CV et les stocker dans le fichier Excel
    note(choix, excel)
    
    # Si le PDF d'origine a été supprimé, ajoutez le nom du fichier à la liste "note_garde"
    if suppres_pdf:
        fichier_sans_dossier = choix.replace(dossier + "/", "")
        note_garde.append(fichier_sans_dossier)


# In[24]:


def suppression_cv(fichier):
    # Déclaration de la variable globale
    global note_garde
    
    # Initialisation d'une variable booléenne "avance" à True
    avance = True
    
    # Parcours de la liste "note_garde"
    for item in note_garde:
        # Vérification si le nom de fichier correspond à un élément de la liste "note_garde"
        if item == fichier:
            # Si une correspondance est trouvée, définissez "avance" sur False et sortez de la boucle
            avance = False
            break
    
    # Si "avance" est toujours True, cela signifie qu'aucune correspondance n'a été trouvée dans la liste "note_garde"
    if avance:
        return
    
    # Construisez le chemin complet du fichier en supprimant l'extension du fichier
    fiche = suppression_extend(fichier)
    
    # Appelez la fonction "suppression_note" pour supprimer les informations associées au fichier
    suppression_note(fiche)


# In[25]:


# Obtenez le chemin du répertoire "Documents" en utilisant la fonction "chemin()"
document = chemin()

# Définissez les chemins pour les répertoires "krecrute", "appel" et "candidat"
krecrute = document + "/Krecrute"
appel = document + "/Krecrute/appel"
candidat = document + "/Krecrute/candidat"

# Créez les répertoires s'ils n'existent pas déjà en utilisant la fonction "creation()"
creation(krecrute)
creation(appel)
creation(candidat)

# Créez un fichier Excel "krecrute.xlsx" et définissez son chemin complet dans la variable "excel"
creation_xlsx("krecrute.xlsx")
excel = krecrute + "/krecrute.xlsx"

# Initialisez des variables globales "suppres_pdf" et "verif" à False et True respectivement
suppres_pdf = False
verif = True

# Initialisez des listes pour stocker les fichiers et dossiers initiaux
init_appel = []
init_dossier = []
init_candidat = {}

# Initialisez des listes pour stocker les fichiers ajoutés et supprimés
fichiers_ajoutes = []
fichiers_supprimes = []

# Initialisez une liste pour stocker les noms de fichiers pour lesquels les notes ont été extraites
note_garde = []

# Boucle principale
while True:
    try:
        # Attendre pendant 5 secondes
        time.sleep(5)
        
        # Appel à la fonction "verification_instance()" pour gérer les répertoires et fichiers
        verification_instance()

        # Obtenir la liste de fichiers actuelle dans le répertoire "appel"
        fichiers_apres = obtenir_fichiers_dans_dossier(appel)

        # Comparer les listes pour détecter les ajouts et suppressions
        fichiers_ajoutes = [fichier for fichier in fichiers_apres if fichier not in init_appel]
        fichiers_supprimes = [fichier for fichier in init_appel if fichier not in fichiers_apres]

        # Mettez à jour la liste initiale des fichiers dans le répertoire "appel"
        init_appel = fichiers_apres

        # Traitez les fichiers ajoutés et supprimés dans le répertoire "appel"
        for fichier in fichiers_ajoutes:
            ajout_appel(fichier)

        for fichier in fichiers_supprimes:
            suppression_appel(fichier)

        # Réinitialisez les listes des fichiers ajoutés et supprimés
        fichiers_ajoutes = []
        fichiers_supprimes = []

        # Parcourez les dossiers dans le répertoire "candidat" et effectuez des opérations similaires
        for cle, valeur in init_candidat.items():
            dossier = candidat + "/" + cle
            fichiers_apres = obtenir_fichiers_dans_dossier(dossier)

            fichiers_ajoutes = [fichier for fichier in fichiers_apres if fichier not in init_candidat[cle]]
            fichiers_supprimes = [fichier for fichier in init_candidat[cle] if fichier not in fichiers_apres]

            init_candidat[cle] = fichiers_apres

            for fichier in fichiers_ajoutes:
                ajout_cv(fichier)

            for fichier in fichiers_supprimes:
                suppression_cv(fichier)

        # Réinitialisez les listes des fichiers ajoutés et supprimés
        fichiers_ajoutes = []
        fichiers_supprimes = []

        # Appel à la fonction "CoUninitialize()" pour finaliser l'application COM
        pythoncom.CoUninitialize()

    except Exception as e:
        print("Une erreur s'est produite :", str(e))

    # Attendre quelques secondes pour permettre d'autres modifications
    time.sleep(5)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




