"""
:author KAPCHE TEBU BAUDOUIN
:version 1.0
"""

import os  # Module pour interagir avec le système de fichiers et les chemins d'accès.
import main
import openpyxl  # Module openpyxl pour la manipulation de fichiers Excel.
import pandas as pd  # Module pour la manipulation et l'analyse de données tabulaires.
from openpyxl.utils.dataframe import dataframe_to_rows  # Fonction pour convertir des données de DataFrame en lignes Excel.
import extraction
import suppression
import re  # Module pour les expressions régulières (regex).
import insertion

def creation_dossier(dossier):
    """
    Cette fonction crée un dossier s'il n'existe pas déjà.

    :param dossier: Le chemin du dossier à créer.
    :type dossier: str
    """
    if not os.path.exists(dossier):
        os.makedirs(dossier)
        print(f"Le dossier '{dossier}' a été créé.")

def creation_excel(nom):
    """
    Cette fonction crée un nouveau fichier Excel avec des feuilles de travail spécifiques.

    :param nom: Le nom du fichier Excel à créer.
    :type nom: str
    """
    excel = main.krecrute + "/" + nom  # Définir le chemin complet du nouveau fichier Excel
    
    if not os.path.exists(excel):
        # Créez un nouveau classeur Excel
        classeur = openpyxl.Workbook()
        
        # Supprimez la feuille de travail par défaut "Sheet"
        nom_de_la_feuille_a_supprimer = "Sheet"
        if nom_de_la_feuille_a_supprimer in classeur.sheetnames:
            feuille_a_supprimer = classeur[nom_de_la_feuille_a_supprimer]
            classeur.remove(feuille_a_supprimer)

        # Créez les feuilles de travail requises
        feuille_projet = classeur.create_sheet("projet")
        feuille_competence = classeur.create_sheet("competence")
        feuille_notation = classeur.create_sheet("notation")
        feuille_metier = classeur.create_sheet("metier")
        feuille_contact = classeur.create_sheet("contact")
        
        # Définissez des données pour les compétences et les projets (ces données sont fournies dans le code)
        # Liste des compétences
        competences=["java","angular","scala","spark","spring","hadoop","postgresql","jenkins","kubernetes","ansible","java EE",
                     "Architecture microservice","spring boot","react","react js","html","git","gitlab","maven","virtualbox",
                     "vagrant","docker","tls","mongodb","sonar","junit","protractor","java spring boot","wildfly","weblogic",
                     "angular js","css","kafka","helm","gitlab ci","cdi","jax-rs","jms","shell","oracle","systeme unix","rhel",
                     "pki","protocol tls","grafana","influxDB","prometheus","collectd","telegraf","appDymanics","jmxtrans",
                     "micrometer","sql devellopper","soapui","rest","telco","cosem","lambda expressions","EJB Runtimes",
                     "terraforme","spring core","JPA","intel IJ","cucumber","kalka","kotlin","even driven architecture",
                     "postgres sql","freeraduis","jdfc","logflash","elasticsearch","telegral","filebeat","spring cloud stream",
                     "linux"]
        
        # Dictionnaire des projets
        projet={'pole':["NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US",
                        "NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US",
                        "NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US","NEX'US"],
                'responsable de pole':["Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali","Eric Mercandali","Eric Mercandali","Eric Mercandali",
                                       "Eric Mercandali"],
                'departement':["SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC","SCC",
                               "CCMM","CCMM","CCMM","CCMM","CCMM","CCMM","CMMA","CMMA","CMMA","CMMA","CMMA","CMMA","SEAL",
                               "SEAL","SEAL","SEAL","SEAL"],
                'responsable de departement':["Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon","Nicolas Pigeon",
                                              "Cedric GAUDET","Cedric GAUDET","Cedric GAUDET","Cedric GAUDET","Cedric GAUDET",
                                              "Cedric GAUDET","pierre GOTELAERE","pierre GOTELAERE","pierre GOTELAERE",
                                              "pierre GOTELAERE","pierre GOTELAERE","pierre GOTELAERE","Eric PERRIER",
                                              "Eric PERRIER","Eric PERRIER","Eric PERRIER","Eric PERRIER"],
                'projet':["ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","ARGOS","mobilite electricque",
                          "mobilite electricque","mobilite electricque","mobilite electricque","MAGELLAN","Gestion Clients",
                          "Gestion Clients","Gestion Clients","usine de Dév","RTE","LINKY ET MAGELLAN","recette transverse",
                          "POA","BEM","deploiment IP","techonologie 2G/3G","PACCMAN","PACCMAN","PACCMAN","iMAT","Expertises",
                          "Expertises","Expertises","Expertises","Cybersecurité"],
                'responsable de projet':["Jean-Bernard Lepidi","Jean-Bernard Lepidi","Jean-Bernard Lepidi",
                                         "Jean-Bernard Lepidi","Jean-Bernard Lepidi","Jean-Bernard Lepidi",
                                         "Jean-Bernard Lepidi","Jean-Bernard Lepidi","Vincent Delalande","Vincent Delalande",
                                         "Vincent Delalande","Vincent Delalande","Alandji BOUORNAKA","Abdourrahmane CHRIABI",
                                         "Abdourrahmane CHRIABI","Abdourrahmane CHRIABI","Sophie CLAPIER","olivier LEROUX",
                                         "cyril Marcant","chrystelle ilicinkas","jean luc LACHAUD","","","","Vincent Boucheux",
                                         "Vincent Boucheux","Vincent Boucheux","Jean-Marc LEBRETON","Benjamin BRETAULT",
                                         "Benjamin BRETAULT","Benjamin BRETAULT","Benjamin BRETAULT","Marc NAUDIN"],
                'application':["CARTOLINE","CAP'ten","Diag BT","PIX'F","PIX'R","Saturne","Tableau de bord energetique",
                               "Bilan metrologiques","mobilite interne","smart charging","comptages du futur","etudes DATA","",
                               "Appui avant vente","realisation e projets smart metting",
                               "appui a la gestion des contrats des clinets externes","","","","","","","","","ICSS","SAR",
                               "Icoeur","","Nouveau comptage et veille technologique","TELCO WAN","OTC","Appui au pilotage",
                               "SEAL Securite"],
                'autre responsable':["","","","","","","","","","","","","","","","","PASCALE ROBERT","PASCALE ROBERT",
                                     "PASCALE ROBERT","PASCALE ROBERT","PASCALE ROBERT","PASCALE ROBERT",
                                     "Anne-Christine LECLERQ","Anne-Christine LECLERQ","Anne-Christine LECLERQ",
                                     "Anne-Christine LECLERQ","Anne-Christine LECLERQ","Anne-Christine LECLERQ",
                                     "RACHED BOUDAOUA","RACHED BOUDAOUA","RACHED BOUDAOUA","RACHED BOUDAOUA",
                                     "RACHED BOUDAOUA"]
               }
        
        # Dictionnaire des contacts (noms et e-mails)
        contact ={"nom":["Nicolas Pigeon","Cedric GAUDET","pierre GOTELAERE","Eric PERRIER","Jean-Bernard Lepidi",
                         "Vincent Delalande","Alandji BOUORNAKA","Abdourrahmane CHRIABI","Sophie CLAPIER","olivier LEROUX",
                         "cyril Marcant","chrystelle ilicinkas","jean luc LACHAUD","Jean-Marc LEBRETON","Benjamin BRETAULT",
                         "Marc NAUDIN","PASCALE ROBERT","Anne-Christine LECLERQ","RACHED BOUDAOUA","Vincent Boucheux"],
                  "email":["nicolas.pigeon@enedis.fr","cedric.gaudet@enedis.fr","pierre.gotelaire@enedis.fr",
                           "eric.perrier@enedis.fr","jean-bernard.lepidi@enedis.fr","vincent.delalande@enedis.fr",
                           "alandji.bouornaka@enedis.fr","abdourrahmane.chriabi@enedis.fr","sophie.clapier@enedis.fr",
                           "olivier.leroux@enedis.fr","cyril.marcant@enedis.fr","chrystelle.ilicinkas@enedis.fr",
                           "jean.lachaud@enedis.fr","jean-marc.lebreton@enedis.fr","benjamin.bretault@enedis.fr",
                           "marc.naudin@enedis.fr","pascale.robert@enedis.fr","anne-christine.leclerq@enedis.fr",
                          "rached.boudaoua@enedis.fr","vincent.boucheux@enedis.fr"],
                  "contact":["","","","","","","","","","","","","","","","","","","",""]
            }

         # Dictionnaire des données métier
        metier={"pole":[],"departement":[],"projet":[],"application":[],"metier":[],"competence":[],"besoin":[],
                "identifiant":[]}
        
        # Dictionnaire des données de notation
        notation={"pole":[],"departement":[],"projet":[],"application":[],"metier":[],"identifiants":[],"contact":[],"email":[],
                  "note":[]}

        # Créez des DataFrames pandas à partir des données
        dataFrame_projet=pd.DataFrame(projet)
        dataFrame_contact=pd.DataFrame(contact)
        dataFrame_metier=pd.DataFrame(metier)
        dataFrame_notation=pd.DataFrame(notation)

        # Utilisez une boucle pour ajouter les données des compétences à la feuille "competence"
        for index, donnee in enumerate(competences, start=1):
            feuille_competence[f"{'A'}{index}"] = donnee

        # Utilisez dataframe_to_rows pour ajouter les données des DataFrames pandas aux feuilles de travail correspondantes
        for row in dataframe_to_rows(dataFrame_projet, index=False, header=True):
            feuille_projet.append(row)

        for row in dataframe_to_rows(dataFrame_contact, index=False, header=True):
            feuille_contact.append(row)

        for row in dataframe_to_rows(dataFrame_metier, index=False, header=True):
            feuille_metier.append(row)

        for row in dataframe_to_rows(dataFrame_notation, index=False, header=True):
            feuille_notation.append(row)

        classeur.save(excel)
        classeur.close()

def nouvel_fiche(new):
    """
    Cette fonction traite un nouveau document PDF et extrait des informations pertinentes pour un candidat.

    :param new: Le chemin du nouveau document PDF à traiter.
    :type new: str
    """
    appel = extraction.extract_appel(new)

    # Définir les chaînes de recherche pour les sections pertinentes du document
    search_string1 = "EXPRESSION DE BESOIN"
    search_string2 = "CONDITIONS D'EXÉCUTION DE LA PRESTATION"
    
    # Initialiser des variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in appel:
        if found == True:
            text += appel[i]
        if item == search_string1:
            found = True
        if item == search_string2:
            found = False
        i += 1

    # Diviser le texte extrait en éléments de besoins
    besoin_client = text.split(".")
    del besoin_client[-1]
    
    # Définir de nouvelles chaînes de recherche pour d'autres sections pertinentes
    search_string1 = "Consultation"
    search_string2 = "Description du projet :"
    
    # Réinitialiser les variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in appel:
        if found == True:
            text += (appel[i] + "\n")
        if item.find(search_string1) != -1:
            found = True
        if item == search_string2:
            found = False
        i += 1
    
    # Diviser le texte extrait en éléments de profil
    profil_primaire = text.split("\n")
    del profil_primaire[-1]
    del profil_primaire[-1]
    
    # Vérifier et ajuster si nécessaire la première ligne du profil
    if profil_primaire[0].find("Projet") == -1:
        del profil_primaire[0]
    
    # Diviser les éléments de profil en intitulés et descriptifs
    profil_secondaire = []
    for item in profil_primaire:
        profil_secondaire += item.split(":")
    
    intitule_profil = []
    descriptif_profil = []
    i = 0
    for item in profil_secondaire:
        if i % 2 == 0:
            intitule_profil.append(item)
        else:
            descriptif_profil.append(item)
        i += 1

    # Définir de nouvelles chaînes de recherche pour d'autres sections pertinentes
    search_string1 = "Stack technique du projet :"
    search_string2 = "Compétences attendues sur le profil :"
    
    # Réinitialiser les variables pour stocker le texte extrait
    i = 0
    text = ""
    found = False
    
    # Parcourir les lignes du texte extrait
    for item in appel:
        if found == True:
            text += (appel[i] + "\n")
        if item.find(search_string1) != -1:
            found = True
        if item.find(search_string2) != -1:
            found = False
        i += 1

    # Diviser le texte extrait en éléments de stack technique
    stack = text.split("\n")
    del stack[-1]
    del stack[-1]

    # Extraire les compétences de la base de données
    competence_extraite = extraction.extract_competence()

    # Créer une liste pour stocker les compétences du profil
    competence_profil = []
    
    # Comparer les compétences extraites avec le texte du stack technique
    for case in competence_extraite:
        for item in stack:
            if re.search(case, item, re.IGNORECASE):
                if case not in competence_profil:
                    competence_profil.append(case)

    # Supprimer les extensions du nom du fichier
    new = suppression.suppression_extend(new)
    
    # Extraire le nom du fichier en supprimant la partie de l'appel
    nom_dossier = new.replace(appel, "")

    # Ajouter les données extraites au fichier Excel global
    global excel
    insertion.ajout_de_donnes(besoin_client, competence_profil, descriptif_profil, nom_dossier, excel)

    # Créer un répertoire pour le candidat et le projet
    chemin_dossier = main.candidat + "/" + nom_dossier
    creation_dossier(chemin_dossier)

